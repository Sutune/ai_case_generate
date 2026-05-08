"""
用例导出模块 (exporter.py)

功能：
  - 将 Markdown 无序列表导出为 XMind 文件（兼容 XMind 20+）
  - 将 Markdown 表格导出为带样式的 Excel 文件
  - 自动保存 Markdown 源文件到 assets/<用例标题>/markdown/，导出结果到 exports/
  - 同名文件自动追加时间戳避免覆盖

用法（命令行）：
  python3 exporter.py xmind '<用例标题>' '<Markdown内容或文件路径>'
  python3 exporter.py excel '<用例标题>' '<Markdown内容或文件路径>'

用法（Python 调用）：
  from exporter import export_from_markdown
  result = export_from_markdown('xmind', '登录流程测试用例', markdown_text)
"""

import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import xmind
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
ASSETS_DIR = PROJECT_ROOT / "assets"
EXPORT_DIR = PROJECT_ROOT / "exports"
MANIFEST_CONTENT = """<?xml version="1.0" encoding="UTF-8"?>
<manifest xmlns="http://www.xmind.net/manifest/2008">
    <file-entry full-path="content.xml" media-type="application/vnd.xmind.workbook+xml"/>
    <file-entry full-path="Thumbnails/thumbnail.png" media-type="image/png"/>
</manifest>"""
EXCEL_COLUMNS = [
    "项目类型",
    "项目",
    "模块",
    "子模块",
    "功能点",
    "用例标题",
    "前置条件",
    "优先级",
    "测试步骤",
    "期望结果",
    "是否自动化",
    "关联需求",
    "是否准入用例",
    "测试结果",
    "用例作者",
    "备注",
    "附件图片",
]
# 需要自动换行的长文本列
_WRAP_COLUMNS = {"测试步骤", "期望结果", "前置条件", "备注"}
# 固定宽度列（字符数）
_COLUMN_WIDTHS = {
    "项目类型": 12, "项目": 12, "模块": 14, "子模块": 14, "功能点": 16,
    "用例标题": 30, "前置条件": 20, "优先级": 8, "测试步骤": 40,
    "期望结果": 40, "是否自动化": 10, "关联需求": 12, "是否准入用例": 10,
    "测试结果": 10, "用例作者": 10, "备注": 20, "附件图片": 14,
}


class ExportError(Exception):
    pass


def sanitize_filename(name: str, fallback: str = "测试用例") -> str:
    cleaned = "".join("_" if char in '<>:"/\\|?*\n\r\t' else char for char in name).strip()
    cleaned = cleaned.rstrip(".")
    return cleaned or fallback


def ensure_dirs() -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def read_markdown_input(raw_input: str) -> str:
    candidate = Path(raw_input)
    if candidate.exists() and candidate.is_file():
        return candidate.read_text(encoding="utf-8")
    return raw_input


def extract_title_from_tree(markdown_text: str) -> str:
    for line in markdown_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("- "):
            return stripped[2:].strip()
    raise ExportError("XMind Markdown 内容缺少根节点标题。")


def normalize_case_title(case_title: str, export_format: str) -> str:
    title = sanitize_filename(case_title)
    if title.endswith(".md"):
        title = title[:-3]
    if export_format in {"excel", "xmind"} and not title.endswith("测试用例"):
        return f"{title}_测试用例"
    return title


def unique_output_path(directory: Path, stem: str, suffix: str) -> Path:
    candidate = directory / f"{stem}{suffix}"
    if not candidate.exists():
        return candidate

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    candidate = directory / f"{stem}_{timestamp}{suffix}"
    if not candidate.exists():
        return candidate

    index = 1
    while True:
        indexed = directory / f"{stem}_{timestamp}_{index:02d}{suffix}"
        if not indexed.exists():
            return indexed
        index += 1


def save_case_markdown(case_title: str, markdown_text: str, export_format: str) -> Path:
    ensure_dirs()
    # 使用 case_title 构建 assets 下的 markdown 目录
    sanitized_title = sanitize_filename(case_title)
    markdown_dir = ASSETS_DIR / sanitized_title / "markdown"
    markdown_dir.mkdir(parents=True, exist_ok=True)
    
    filename = normalize_case_title(case_title, export_format)
    markdown_stem = f"{filename}_{export_format}"
    output_path = unique_output_path(markdown_dir, markdown_stem, ".md")
    output_path.write_text(markdown_text.strip() + "\n", encoding="utf-8")
    return output_path


# ── XMind 导出 ──────────────────────────────────────────────────────────────

def parse_tree_markdown(markdown_text: str) -> dict:
    lines = [line.rstrip() for line in markdown_text.splitlines() if line.strip()]
    items: list[tuple[int, str]] = []

    for line in lines:
        match = re.match(r"^(\s*)-\s+(.+?)\s*$", line)
        if not match:
            raise ExportError("XMind Markdown 必须全部使用 '-' 无序列表格式。")
        indent = len(match.group(1))
        title = match.group(2).strip()
        items.append((indent, title))

    if not items:
        raise ExportError("XMind Markdown 内容为空。")

    root_indent, root_title = items[0]
    root = {"title": root_title, "children": []}
    stack = [(root_indent, root)]

    for indent, title in items[1:]:
        while stack and indent <= stack[-1][0]:
            stack.pop()
        if not stack:
            raise ExportError("XMind Markdown 层级不合法，请检查缩进。")
        node = {"title": title, "children": []}
        stack[-1][1]["children"].append(node)
        stack.append((indent, node))

    return root


def append_topics(parent_topic, children: list[dict]) -> None:
    for child in children:
        topic = parent_topic.addSubTopic()
        topic.setTitle(child["title"])
        append_topics(topic, child["children"])


def repair_xmind_for_xmind20(path: Path) -> None:
    if not zipfile.is_zipfile(path):
        raise ExportError(f"非法的 XMind 文件: {path}")

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        with zipfile.ZipFile(path, "r") as archive:
            archive.extractall(temp_root)

        # 校验 content.xml 存在，否则 manifest 声明无意义
        if not (temp_root / "content.xml").exists():
            raise ExportError("XMind 文件缺少 content.xml，无法完成兼容性修复。")

        meta_inf = temp_root / "META-INF"
        meta_inf.mkdir(parents=True, exist_ok=True)
        (meta_inf / "manifest.xml").write_text(MANIFEST_CONTENT, encoding="utf-8")

        # 使用唯一临时文件名避免并发冲突，写入完成后原子替换
        fd, tmp_str = tempfile.mkstemp(dir=path.parent, suffix=".tmp")
        tmp_path = Path(tmp_str)
        try:
            os.close(fd)
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as archive:
                for file_path in temp_root.rglob("*"):
                    if file_path.is_file():
                        archive.write(file_path, file_path.relative_to(temp_root))
            shutil.move(str(tmp_path), str(path))
        except Exception:
            tmp_path.unlink(missing_ok=True)
            raise


def export_xmind(case_title: str, markdown_text: str) -> tuple[Path, Path]:
    markdown_path = save_case_markdown(case_title, markdown_text, "xmind")
    tree = parse_tree_markdown(markdown_text)
    export_title = normalize_case_title(case_title or tree["title"], "xmind")
    xmind_path = unique_output_path(EXPORT_DIR, export_title, ".xmind")

    workbook = xmind.load(str(xmind_path))
    sheet = workbook.getPrimarySheet()
    sheet.setTitle(tree["title"])
    root_topic = sheet.getRootTopic()
    root_topic.setTitle(tree["title"])
    append_topics(root_topic, tree["children"])

    xmind.save(workbook, path=str(xmind_path))
    repair_xmind_for_xmind20(xmind_path)
    return markdown_path, xmind_path


# ── Excel 导出 ──────────────────────────────────────────────────────────────

def _normalize_header_cell(cell: str) -> str:
    """标准化表头单元格：去除全角/半角空格差异。"""
    return cell.strip().replace("\u3000", "").replace(" ", "")


def split_markdown_row(line: str) -> list[str]:
    if not line.strip().startswith("|"):
        raise ExportError("Excel Markdown 必须使用标准表格格式。")
    cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
    return cells


def parse_table_markdown(markdown_text: str) -> list[dict]:
    lines = [line.rstrip() for line in markdown_text.splitlines() if line.strip()]
    table_lines = [line for line in lines if line.lstrip().startswith("|")]
    if len(table_lines) < 3:
        raise ExportError("Excel Markdown 至少需要表头、分隔行和一行数据。")

    header_raw = split_markdown_row(table_lines[0])
    delimiter = split_markdown_row(table_lines[1])

    # 分隔行校验：非空 cell 必须全为 '-' 和 ':'，空 cell 视为对齐占位（不报错）
    for cell in delimiter:
        if cell and not all(c in "-:" for c in cell):
            raise ExportError(f"Excel Markdown 第二行必须是表格分隔行，发现非法内容: {cell!r}")

    # 表头容错：归一化后比对，给出具体差异提示
    header_normalized = [_normalize_header_cell(h) for h in header_raw]
    expected_normalized = [_normalize_header_cell(col) for col in EXCEL_COLUMNS]
    if header_normalized != expected_normalized:
        missing = [c for c in EXCEL_COLUMNS if _normalize_header_cell(c) not in header_normalized]
        extra = [h for h in header_raw if _normalize_header_cell(h) not in expected_normalized]
        parts = []
        if missing:
            parts.append(f"缺少字段: {missing}")
        if extra:
            parts.append(f"多余字段: {extra}")
        raise ExportError("Excel Markdown 表头与约定字段不一致。" + ("；".join(parts) if parts else ""))

    rows: list[dict] = []
    for line in table_lines[2:]:
        values = split_markdown_row(line)
        if len(values) != len(EXCEL_COLUMNS):
            raise ExportError(
                f"Excel Markdown 数据列数（{len(values)}）与表头列数（{len(EXCEL_COLUMNS)}）不一致。"
            )
        if all(not v for v in values):
            continue
        rows.append(dict(zip(EXCEL_COLUMNS, values)))

    if not rows:
        raise ExportError("Excel Markdown 没有有效数据行。")
    return rows


def _apply_excel_style(excel_path: Path) -> None:
    """为导出的 Excel 设置表头样式、列宽和自动换行。"""
    wb = load_workbook(excel_path)
    ws = wb.active

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[col_letter].width = _COLUMN_WIDTHS.get(col_name, 14)

    wrap_align = Alignment(vertical="top", wrap_text=True)
    normal_align = Alignment(vertical="top", wrap_text=False)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_name = EXCEL_COLUMNS[cell.column - 1] if cell.column <= len(EXCEL_COLUMNS) else ""
            cell.alignment = wrap_align if col_name in _WRAP_COLUMNS else normal_align

    ws.freeze_panes = "A2"
    wb.save(excel_path)


def export_excel(case_title: str, markdown_text: str) -> tuple[Path, Path]:
    markdown_path = save_case_markdown(case_title, markdown_text, "excel")
    rows = parse_table_markdown(markdown_text)
    export_title = normalize_case_title(case_title, "excel")
    excel_path = unique_output_path(EXPORT_DIR, export_title, ".xlsx")
    dataframe = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    dataframe.to_excel(excel_path, index=False)
    _apply_excel_style(excel_path)
    return markdown_path, excel_path


# ── 入口 ────────────────────────────────────────────────────────────────────

def export_from_markdown(export_format: str, case_title: str, markdown_input: str) -> dict:
    markdown_text = read_markdown_input(markdown_input)
    if export_format == "xmind" and not case_title:
        case_title = extract_title_from_tree(markdown_text)
    if export_format == "excel" and not case_title:
        case_title = "测试用例"

    if export_format == "xmind":
        markdown_path, export_path = export_xmind(case_title, markdown_text)
    elif export_format == "excel":
        markdown_path, export_path = export_excel(case_title, markdown_text)
    else:
        raise ExportError(f"不支持的导出格式: {export_format}")

    return {
        "status": "success",
        "markdown_path": str(markdown_path),
        "export_path": str(export_path),
        "format": export_format,
    }


def main() -> None:
    if len(sys.argv) < 4:
        print(
            json.dumps(
                {"error": "Usage: python exporter.py <excel|xmind> <case_title> <markdown_or_markdown_path>"},
                ensure_ascii=False,
            )
        )
        sys.exit(1)

    export_format = sys.argv[1].lower().strip()
    case_title = sys.argv[2].strip()
    markdown_input = sys.argv[3]

    try:
        result = export_from_markdown(export_format, case_title, markdown_input)
        print(json.dumps(result, ensure_ascii=False))
    except ExportError as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        sys.exit(1)
    except Exception as exc:
        print(json.dumps({"error": f"导出失败: {exc}"}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
